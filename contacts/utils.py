# youbee_contacts/contacts/utils.py

# System libraries
import os

# Third-party libraries
from openpyxl import Workbook

# Django modules
from django.db.models import Value, CharField
from django.db.models.functions import Concat
from django.db.models import Func, F
from django.conf import settings

# Django apps
from youbee.storage_backends import media_storage

# Current-app modules
from .models import Contact


class Levenshtein(Func):
    template = "%(function)s(%(expressions)s, '%(search_term)s')"
    function = "levenshtein"

    def __init__(self, expression, search_term, ins_cost=1, del_cost=1, sub_cost=1, **extras):
        super(Levenshtein, self).__init__(
            expression,
            search_term=search_term,
            ins_cost=ins_cost,
            del_cost=del_cost,
            sub_cost=sub_cost,
            **extras
        )


def get_contacts_annotated_with_complete_name(company):
    q = Contact.objects.filter(company=company) \
        .annotate(complete_name=Concat('last_name', Value(' '), 'first_name', output_field=CharField()))\
        .only('first_name', 'last_name', 'function', 'organisation', 'email1')

    return q


def get_similar_contacts(company, name):
    q = get_contacts_annotated_with_complete_name(company=company)

    return q.annotate(lev_dist=Levenshtein(F('complete_name'), name)).filter(lev_dist__lte=5)\
        .order_by('lev_dist')


def export_contact_to_excel(contact_list):
    c = contact_list.first()
    filepath = os.path.join(settings.MEDIA_ROOT, settings.EXPORT_DIR, settings.COMPANY_DIR + str(c['company_id']), 'export_contacts.xlsx')
    filename = media_storage.generate_filename(filepath)
    wb = Workbook()
    sheet = wb.create_sheet('Contacts')

    #  Create the headings row
    headings = ['title', 'first_name', 'last_name', 'function', 'organisation', 'gender', 'partner', 'email1', 'email2']
    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    #  Create a row for each contact in the query
    for rowno, row in enumerate(contact_list, start=2):
        for colno, field_name in enumerate(headings, start=1):
            print(row)
            sheet.cell(row=rowno, column=colno).value = row[field_name]

    wb.save(filename)

    return filename, media_storage.exists(filename)
